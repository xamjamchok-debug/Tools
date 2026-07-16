"""Gegenbuchungs-Automatik + Haushalts-Saldo-Korrektur.

Zwei Aufgaben, in einem Lauf:

1. korrigiere_stammdaten() — Rollen/Flags + private Beträge (aus lokaler FB):
     * Kategorie-Rolle `zaehlt_als`: 'ruecklage' (Topf, wird abgezogen) | 'forderung'
       (Natalie/Jörg, wird addiert) | 'ausgabe' (Haushaltskasse/Einnahmen, kein Topf).
     * Vermögensposten: ETF-Wert fixen (7.000 statt 70.000), Merkzettel (−6.800) und
       Großeltern-geparkt (−72.975) anlegen — alle mit im_haushaltssaldo=TRUE.
       Kredit Großeltern/Riester/KfW/Deutsche-Bank -> im_haushaltssaldo=FALSE (langfristig).
     * Autohaus-Meures-Buchungen als 'kreditfinanziert' markieren -> kein Rücklagen-Verzehr.

2. sync_gegenbuchungen() — für Input-Realbuchungen (quelle NICHT aus der FB) mit einer
   Rücklagen-Kategorie je eine verknüpfte Spiegel-Buchung (Verzehr/Zuführung) erzeugen.
   Die Spiegelung hängt via spiegel_von_id an der Realbuchung und folgt ihrer Kategorie
   (Umbuchen jederzeit möglich). Idempotent. Kreditfinanzierte + Nicht-Topf-Kategorien
   ('ausgabe'/'forderung') bekommen KEINE Spiegelung.

Beträge kommen aus der lokalen Fuchsbaukasse, NIE aus dem Code (Datenschutz).

Aufruf:
  python -m haushaltskasse.workflows.gegenbuchung           # Vorschau, kein Schreiben
  python -m haushaltskasse.workflows.gegenbuchung --write    # schreiben
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from ..domain.saldo import FB_QUELLEN, SQL_SPIEGEL_BERECHTIGT   # #60: eine Quelle der Regel
from ..storage.db import connect, init_db
from .migration_fb import FB_PATH

# --- Kategorie-Rollen (Rest bleibt 'ruecklage') -----------------------------
ROLLE_FORDERUNG = ("Natalie", "Jörg")
# Haushaltskasse ist jetzt ein EIGENER Rücklagen-Topf (Kategorie „0") → NICHT mehr hier.
# 'Einnahmen' bleibt ohne Topf (#46); 'Kinder' entfällt mit dem Cleanup (#49).
ROLLE_AUSGABE = ("Einnahmen", "Kinder")

# --- Vermögensposten, die NICHT in den Haushalts-Saldo zählen (langfristig) --
POSTEN_LANGFRIST = ("Kredit an Großeltern", "Riester-Steuerschuld",
                    "KfW-Kredit", "Deutsche-Bank-Kredit")

KREDITFINANZIERT = "kreditfinanziert (Großeltern-Darlehen)"
AUTOHAUS_MUSTER = "%Autohaus Meures%"


# ---------------------------------------------------------------------------
# Fuchsbaukasse-„Übersicht" auslesen (linke + rechte Spalte), nur Beträge.
# ---------------------------------------------------------------------------
def _fb_uebersicht() -> tuple[dict[str, int], dict[str, int]]:
    tmp = Path(tempfile.gettempdir()) / "gb_fb.xlsm"
    shutil.copy2(FB_PATH, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True, keep_vba=False)
    ws = wb[next(s for s in wb.sheetnames if s.replace("﻿", "").endswith("bersicht"))]

    def _cent(x):
        try:
            return round(float(x) * 100)
        except (ValueError, TypeError):
            return None

    links: dict[str, int] = {}
    rechts: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        if len(r) > 3 and r[2] and _cent(r[3]) is not None:
            links[str(r[2]).strip()] = _cent(r[3])
        if len(r) > 8 and r[7] and _cent(r[8]) is not None:
            rechts[str(r[7]).strip()] = _cent(r[8])
    wb.close()
    return links, rechts


# ---------------------------------------------------------------------------
# 1) Stammdaten korrigieren
# ---------------------------------------------------------------------------
def korrigiere_stammdaten(cur) -> list[str]:
    links, rechts = _fb_uebersicht()
    log: list[str] = []

    # Kategorie-Rollen
    cur.execute("UPDATE kategorien SET zaehlt_als='ruecklage'")
    cur.execute("UPDATE kategorien SET zaehlt_als='forderung' WHERE name = ANY(%s)", (list(ROLLE_FORDERUNG),))
    log.append(f"forderung-Rolle: {cur.rowcount} Kategorien (Natalie/Jörg)")
    cur.execute("UPDATE kategorien SET zaehlt_als='ausgabe' WHERE name = ANY(%s)", (list(ROLLE_AUSGABE),))
    log.append(f"ausgabe-Rolle:   {cur.rowcount} Kategorien (Haushaltskasse/Einnahmen/Kinder)")

    # ETF-Wert korrigieren (aus FB) + sicher im Saldo
    if "ETFs" in links:
        cur.execute("""UPDATE vermoegensposten SET wert_cent=%s, art='vermoegen', im_haushaltssaldo=TRUE
                       WHERE name='ETFs / Depot'""", (links["ETFs"],))
        log.append(f"ETF-Posten -> {links['ETFs']/100:,.2f} € (Fix, war 70.000)")

    # Merkzettel + Anlage Großeltern als Saldo-relevante Posten (Beträge aus FB)
    _upsert_posten(cur, "Merkzettel", links.get("Merkzettel", 0), "vermoegen", True,
                   "Merkliste offener Posten (FB-Übersicht)")
    log.append(f"Merkzettel-Posten -> {links.get('Merkzettel',0)/100:,.2f} €")
    geparkt = -abs(rechts.get("Comdirect", 0))   # geliehenes Geld der Großeltern -> Abzug
    _upsert_posten(cur, "Anlage Großeltern", geparkt, "vermoegen", True,
                   "geliehenes Geld der Großeltern, geparkt auf comdirect-Tagesgeld")
    log.append(f"Anlage-Großeltern-Posten -> {geparkt/100:,.2f} €")

    # Langfrist-Posten: aus dem Haushalts-Saldo nehmen und (falls früher soft-gelöscht,
    # um sie aus dem alten Saldo zu halten) reaktivieren — das Flag trennt jetzt sauber.
    cur.execute("UPDATE vermoegensposten SET im_haushaltssaldo=FALSE, aktiv=TRUE WHERE name = ANY(%s)",
                (list(POSTEN_LANGFRIST),))
    log.append(f"langfristig (nicht im Saldo, reaktiviert): {cur.rowcount} Posten")

    # Autohaus-Meures = kreditfinanziert (kein Rücklagen-Verzehr)
    cur.execute("""UPDATE buchungen SET bemerkung=%s
                   WHERE buchungsart='real' AND quelle_import <> ALL(%s)
                     AND empfaenger ILIKE %s""",
                (KREDITFINANZIERT, list(FB_QUELLEN), AUTOHAUS_MUSTER))
    log.append(f"kreditfinanziert markiert: {cur.rowcount} Autohaus-Buchungen")
    return log


def _upsert_posten(cur, name, wert_cent, art, im_saldo, notiz) -> None:
    cur.execute("""INSERT INTO vermoegensposten (name, wert_cent, art, im_haushaltssaldo, notiz)
                   VALUES (%s,%s,%s,%s,%s)
                   ON CONFLICT (name) DO UPDATE
                   SET wert_cent=EXCLUDED.wert_cent, art=EXCLUDED.art,
                       im_haushaltssaldo=EXCLUDED.im_haushaltssaldo, notiz=EXCLUDED.notiz""",
                (name, wert_cent, art, im_saldo, notiz))


# ---------------------------------------------------------------------------
# 2) Gegenbuchungen synchronisieren
# ---------------------------------------------------------------------------
def _spiegel_upsert(cur, real_id, datum, betrag, kat_id, ukat_id) -> None:
    cur.execute("""
        INSERT INTO buchungen (buchungsart, datum_wert, betrag_cent, kategorie_id,
                               unterkategorie_id, spiegel_von_id, quelle_import,
                               import_hash, bemerkung)
        VALUES ('ruecklage',%s,%s,%s,%s,%s,'spiegel',%s,'Auto-Gegenbuchung (Verzehr/Zuführung)')
        ON CONFLICT (import_hash) DO UPDATE
        SET betrag_cent=EXCLUDED.betrag_cent, datum_wert=EXCLUDED.datum_wert,
            kategorie_id=EXCLUDED.kategorie_id, unterkategorie_id=EXCLUDED.unterkategorie_id""",
        (datum, betrag, kat_id, ukat_id, real_id, f"spiegel-{real_id}"))


def sync_eine(cur, real_id: int) -> None:
    """Bringt die Spiegel-Gegenbuchung EINER Realbuchung mit ihrer aktuellen Kategorie in
    Einklang. Wird beim Umkategorisieren im Dashboard aufgerufen -> Gegenbuchung folgt der
    Kategorie: falsch zugeordnet einfach umkategorisieren, der Spiegel wandert automatisch mit.
    """
    cur.execute("DELETE FROM buchungen WHERE spiegel_von_id=%s", (real_id,))
    cur.execute(f"""
        SELECT b.datum_wert, b.betrag_cent, b.kategorie_id, b.unterkategorie_id
        FROM buchungen b JOIN kategorien k ON k.id=b.kategorie_id
        WHERE b.id=%s AND {SQL_SPIEGEL_BERECHTIGT}""",
        (real_id, list(FB_QUELLEN)))
    row = cur.fetchone()
    if row:
        _spiegel_upsert(cur, real_id, *row)


def sync_gegenbuchungen(cur) -> dict:
    """Erzeugt/aktualisiert je berechtigter Realbuchung eine verknüpfte Spiegel-Buchung.
    Berechtigt = Input-Realbuchung (nicht FB), Kategorie mit zaehlt_als='ruecklage',
    nicht kreditfinanziert. Gibt Statistik zurück."""
    # verwaiste Spiegel entfernen (Quelle nicht mehr berechtigt -> z. B. umkategorisiert)
    cur.execute(f"""
        DELETE FROM buchungen s
        WHERE s.spiegel_von_id IS NOT NULL
          AND NOT EXISTS (
            SELECT 1 FROM buchungen b JOIN kategorien k ON k.id=b.kategorie_id
            WHERE b.id = s.spiegel_von_id AND {SQL_SPIEGEL_BERECHTIGT}
          )""", (list(FB_QUELLEN),))
    geloescht = cur.rowcount

    # fehlende/veraltete Spiegel anlegen bzw. aktualisieren (Betrag/Kategorie folgen der Quelle)
    cur.execute(f"""
        SELECT b.id, b.datum_wert, b.betrag_cent, b.kategorie_id, b.unterkategorie_id
        FROM buchungen b JOIN kategorien k ON k.id=b.kategorie_id
        WHERE {SQL_SPIEGEL_BERECHTIGT}
        ORDER BY b.id""", (list(FB_QUELLEN),))
    quellen = cur.fetchall()
    for bid, datum, betrag, kat_id, ukat_id in quellen:
        _spiegel_upsert(cur, bid, datum, betrag, kat_id, ukat_id)
    return {"quellen": len(quellen), "geloescht": geloescht}


# ---------------------------------------------------------------------------
# Bericht: Haushalts-Saldo nach neuer Formel
# ---------------------------------------------------------------------------
def _saldo_bericht(cur) -> None:
    from ..dashboard.queries import haushaltssaldo
    s = haushaltssaldo(cur)
    e = lambda c: (c or 0) / 100
    print("\n--- Haushalts-Saldo (neue Formel) ---")
    print(f"  Konten (real)                {e(s['konten_cent']):>14,.2f}")
    print(f"  + Posten im Saldo            {e(s['posten_cent']):>14,.2f}   (ETF, Merkzettel, Großeltern-geparkt)")
    print(f"  + Forderungen (Natalie/Jörg) {e(s['forderung_cent']):>14,.2f}")
    print(f"  - Ruecklagen-Toepfe          {e(s['ruecklagen_cent']):>14,.2f}")
    print(f"  {'='*44}")
    print(f"  = Haushalts-Saldo            {e(s['saldo_cent']):>14,.2f}   (FB-Ziel: -14.579 vor Amazon)")
    print(f"\n  Langfristig (separat, NICHT im Saldo): {e(s['langfrist_cent']):,.2f}")


def lauf(write: bool = False) -> None:
    from .audit import kennzahlen_json, protokolliere   # #61: Audit + Invarianten-Abschluss
    init_db()   # legt die neuen Spalten idempotent an
    conn = connect()
    try:
        vorher = kennzahlen_json(conn) if write else ""
        with conn.cursor() as cur:
            print("[gegenbuchung] Stammdaten korrigieren ...")
            for zeile in korrigiere_stammdaten(cur):
                print(f"  - {zeile}")
            print("\n[gegenbuchung] Gegenbuchungen synchronisieren ...")
            stat = sync_gegenbuchungen(cur)
            print(f"  - {stat['quellen']} Spiegel-Buchungen aktiv, {stat['geloescht']} verwaiste entfernt")
            _saldo_bericht(cur)
        if write:
            ok = protokolliere(conn, "gegenbuchung", "--write", vorher)
            conn.commit()
            print(f"\n[gegenbuchung] committed. Invarianten: {'OK' if ok else 'VERLETZT (s. oben)'}\n")
        else:
            conn.rollback()
            print("\n[Vorschau] Rollback – nichts geschrieben. Mit --write ausführen.\n")
    finally:
        conn.close()


if __name__ == "__main__":
    lauf(write="--write" in sys.argv)
