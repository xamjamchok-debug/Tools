"""Einmal-Migration der alten Excel-Fuchsbaukasse in die Postgres-DB.

Liest Fuchsbaukasse2026_claude.xlsm (über eine Kopie, um den Excel-Lock zu umgehen):
  * DKB-Blatt        -> buchungen (buchungsart='real',      Konto DKB-Giro)
  * Kto-*-Blätter    -> buchungen (buchungsart='ruecklage', je Kategorie)

Verifikation: Summe je Konto/Kategorie muss den Salden der Übersicht entsprechen.

Aufruf:  python -m haushaltskasse.workflows.migration_fb          (Vorschau + Verifikation, KEIN Schreiben)
         python -m haushaltskasse.workflows.migration_fb --write  (TRUNCATE buchungen + Migration schreiben)
"""
from __future__ import annotations

import hashlib
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from ..storage.db import connect

FB_PATH = Path(__file__).resolve().parent.parent / "input" / "Fuchsbaukasse2026_claude.xlsm"

# Kto-Blatt (nach 'Kto-') -> kanonische Kategorie (= Nebenbuch-Name im seed)
def _kategorie_fuer(blatt: str) -> str | None:
    rest = blatt[4:]  # nach 'Kto-'
    fest = {
        "Auto": "Auto", "Sport": "Sport", "Urlaub": "Urlaub", "NK": "Nebenkosten",
        "Tel": "Telefon", "TK": "TK", "Vers": "Vers", "Kredit": "Kredit",
        "Inst": "Inst", "Natalie": "Natalie",
    }
    if rest in fest:
        return fest[rest]
    if rest.startswith("F"):   # Füchschen (Umlaut-sicher)
        return "Füchschen"
    if rest.startswith("J"):   # Jörg
        return "Jörg"
    return None


def _excel_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and 20000 < v < 60000:   # Excel-Seriennummer
        return date(1899, 12, 30) + timedelta(days=int(v))
    if isinstance(v, str):                                    # Text-Datum, z. B. '04.06.2018'
        s = v.strip()
        for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _cent(v) -> int | None:
    if isinstance(v, (int, float)):
        return round(float(v) * 100)
    return None


def _hash(*teile) -> str:
    return hashlib.sha1("|".join(str(t) for t in teile).encode("utf-8")).hexdigest()[:16]


def parse_dkb_sheet(ws) -> list[dict]:
    """DKB-Blatt -> reale Buchungen. Spalten: Buchungstag(0) Wertstellung(1) Buchungstext(2)
    Auftraggeber(3) Empfänger(4) Verwendungszweck(5) Typ(6) IBAN(7) Betrag(8)."""
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # Header
        if len(r) < 9:
            continue
        betrag = _cent(r[8])
        datum = _excel_date(r[1]) or _excel_date(r[0])
        if betrag is None or datum is None:
            continue
        empf = str(r[4] or r[3] or "").strip()
        zweck = str(r[5] or "").strip()
        vorgang = str(r[6] or "").strip()   # Typ (z. B. 'Überweisung', 'Kartenzahlung')
        iban = str(r[7] or "").strip()      # Gegen-IBAN
        out.append({
            "datum": datum.isoformat(),
            "betrag_cent": betrag,
            "empfaenger": empf[:200],
            "verwendungszweck": zweck[:500],
            "vorgang": vorgang[:100],
            "iban_gegen": iban,
            "import_hash": _hash("fb-dkb", i, datum, betrag, empf[:40]),
        })
    return out


def parse_kto_sheet(ws) -> list[dict]:
    """Kto-Blatt -> Rücklagen-Buchungen. Spalten: Datum(0) Betrag(1) Saldo(2) Was(3)
    Referenz(4) Auftraggeber(5) Empfänger(6) Unterkategorie(7) Bemerkung(8) ID(9)."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next((i for i, r in enumerate(rows) if str(r[0]).strip().lower() == "datum"), 0)
    out = []
    for i, r in enumerate(rows[hdr_i + 1:], start=hdr_i + 2):
        if len(r) < 2:
            continue
        betrag = _cent(r[1])
        datum = _excel_date(r[0])
        if betrag is None or datum is None:
            continue
        was = str(r[3] or "").strip() if len(r) > 3 else ""
        ukat = str(r[7] or "").strip() if len(r) > 7 else ""
        bem = str(r[8] or "").strip() if len(r) > 8 else ""
        out.append({
            "datum": datum.isoformat(),
            "betrag_cent": betrag,
            "verwendungszweck": was[:500],
            "unterkategorie": ukat[:100],
            "bemerkung": bem[:200],
            "import_hash": _hash("fb-kto", ws.title, i, datum, betrag, was[:40]),
        })
    return out


def _lade_fuchsbaukasse() -> tuple[list[dict], dict[str, list[dict]]]:
    tmp = Path(tempfile.gettempdir()) / "fb_migrate_copy.xlsm"
    shutil.copy2(FB_PATH, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True, keep_vba=False)
    dkb = parse_dkb_sheet(wb["DKB"])
    kto: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        if name.startswith("Kto-"):
            kat = _kategorie_fuer(name)
            if kat:
                kto[kat] = parse_kto_sheet(wb[name])
    wb.close()
    return dkb, kto


def _verifikation(dkb: list[dict], kto: dict[str, list[dict]]) -> None:
    print("\n--- Verifikation (Summen) ---")
    s = sum(b["betrag_cent"] for b in dkb) / 100
    print(f"  DKB-Giro (real): {len(dkb):5d} Buchungen  Summe {s:>12,.2f} €   (Soll: FB-Übersichtsblatt)")
    gesamt_rueck = 0
    for kat, buchungen in sorted(kto.items()):
        s = sum(b["betrag_cent"] for b in buchungen)
        gesamt_rueck += s
        print(f"  {kat:14s} (ruecklage): {len(buchungen):5d}  Summe {s/100:>11,.2f} €")
    print(f"  {'Summe Rücklagen':14s}: {gesamt_rueck/100:>28,.2f} €")


def migriere(write: bool = False) -> None:
    dkb, kto = _lade_fuchsbaukasse()
    _verifikation(dkb, kto)
    if not write:
        print("\n[Vorschau] Kein Schreiben. Mit --write ausführen, um in die DB zu migrieren.")
        return

    conn = connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT name, id FROM konten")
            konten = dict(cur.fetchall())
            cur.execute("SELECT name, id FROM kategorien")
            kategorien = dict(cur.fetchall())

            print("\n[migrate] TRUNCATE buchungen (Neuaufbau) ...")
            cur.execute("TRUNCATE buchungen RESTART IDENTITY CASCADE")

            # DKB real
            dkb_id = konten["DKB-Giro"]
            for b in dkb:
                cur.execute(
                    "INSERT INTO buchungen (buchungsart, datum_wert, betrag_cent, konto_id, "
                    "empfaenger, verwendungszweck, quelle_import, import_hash) "
                    "VALUES ('real',%s,%s,%s,%s,%s,'fb-dkb',%s) ON CONFLICT (import_hash) DO NOTHING",
                    (b["datum"], b["betrag_cent"], dkb_id, b["empfaenger"],
                     b["verwendungszweck"], b["import_hash"]),
                )

            # Kto ruecklage
            ukat_cache: dict = {}
            for kat, buchungen in kto.items():
                kat_id = kategorien.get(kat)
                if kat_id is None:
                    print(f"  [WARN] Kategorie '{kat}' nicht in DB – übersprungen.")
                    continue
                for b in buchungen:
                    ukat_id = None
                    uname = b.get("unterkategorie")
                    if uname:
                        key = (kat_id, uname)
                        if key not in ukat_cache:
                            cur.execute(
                                "INSERT INTO unterkategorien (kategorie_id, name, quelle) "
                                "VALUES (%s,%s,'manuell') ON CONFLICT (kategorie_id, name) DO NOTHING",
                                (kat_id, uname))
                            cur.execute("SELECT id FROM unterkategorien WHERE kategorie_id=%s AND name=%s",
                                        (kat_id, uname))
                            ukat_cache[key] = cur.fetchone()[0]
                        ukat_id = ukat_cache[key]
                    cur.execute(
                        "INSERT INTO buchungen (buchungsart, datum_wert, betrag_cent, kategorie_id, "
                        "unterkategorie_id, verwendungszweck, bemerkung, quelle_import, import_hash) "
                        "VALUES ('ruecklage',%s,%s,%s,%s,%s,%s,'fb-kto',%s) ON CONFLICT (import_hash) DO NOTHING",
                        (b["datum"], b["betrag_cent"], kat_id, ukat_id,
                         b["verwendungszweck"], b["bemerkung"], b["import_hash"]),
                    )
        conn.commit()

        # Kontrolle aus der DB
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM buchungen")
            n = cur.fetchone()[0]
            cur.execute("SELECT k.name, COALESCE(SUM(b.betrag_cent),0) FROM buchungen b "
                        "JOIN konten k ON k.id=b.konto_id WHERE b.buchungsart='real' GROUP BY k.name")
            real = cur.fetchall()
            cur.execute("SELECT COALESCE(SUM(betrag_cent),0) FROM buchungen WHERE buchungsart='ruecklage'")
            rueck = cur.fetchone()[0]
        print(f"\n[migrate] geschrieben: {n} Buchungen gesamt.")
        for name, s in real:
            print(f"  real {name}: {s/100:,.2f} €")
        print(f"  Summe Rücklagen (DB): {rueck/100:,.2f} €")
    finally:
        conn.close()


if __name__ == "__main__":
    migriere(write="--write" in sys.argv)
