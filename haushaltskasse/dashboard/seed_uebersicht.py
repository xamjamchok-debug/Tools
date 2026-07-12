"""Einmal-Seed für Übersicht & Rücklagen aus der Fuchsbaukasse.

  * config-Blatt  -> monatliche Soll-Rückstellung je Kategorie/Unterkategorie
  * Übersicht/Kredit_Großeltern -> Vermögensposten (ETFs, Kredit an Großeltern,
    Riester-Steuerschuld, KfW-/Deutsche-Bank-Kredit als Platzhalter).

Alle Werte sind danach im Dashboard editierbar (DB = Quelle der Wahrheit). Idempotent.

Aufruf:  python -m haushaltskasse.dashboard.seed_uebersicht
"""
from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from ..storage.db import connect
from ..workflows.migration_fb import FB_PATH

# config-Name -> Kategorie (Soll auf Kategorie-Ebene)
CONFIG_KATEGORIE = {
    "Kredite": "Kredit", "Urlaub": "Urlaub", "Auto": "Auto", "Sport": "Sport",
    "Telefon": "Telefon", "Nebenkosten": "Nebenkosten", "Inst": "Inst",
    "Haushaltskasse": "Haushaltskasse", "Jörg": "Jörg", "Natalie": "Natalie",
    "Füchschen": "Füchschen", "Krankenkasse": "TK",
}
# config-Name -> (Kategorie, Unterkategorie)  (Soll auf Unterkategorie-Ebene)
CONFIG_UNTERKATEGORIE = {
    "Haftpflicht": ("Vers", "Haftpflicht"),
    "Strom": ("Nebenkosten", "Strom"),
}

# Externe Posten: (Label im FB-„Übersicht"-Blatt | None, DB-Name, art, notiz).
# Die BETRÄGE werden aus der lokalen Fuchsbaukasse gelesen (bleiben aus Datenschutzgründen
# NICHT im Code/Repo). Posten mit Label=None werden als 0-Platzhalter angelegt und im Dashboard gefüllt.
POSTEN_QUELLEN = [
    ("ETF",               "ETFs / Depot",          "vermoegen", "Marktwert Wertpapierdepot"),
    ("Riester",           "Riester-Steuerschuld",  "schuld",    None),
    ("Kredit Großeltern", "Kredit an Großeltern",  "schuld",    "Am Königsberg"),
    (None,                "KfW-Kredit",            "schuld",    "Restschuld eintragen"),
    (None,                "Deutsche-Bank-Kredit",  "schuld",    "Restschuld eintragen"),
]


def _uebersicht_werte() -> dict[str, int]:
    """Liest {Label: Cent} aus dem FB-„Übersicht"-Blatt (Label in Spalte C, Wert in Spalte D).
    Robust gegen Text/Zahl und #REF!-Fehler."""
    tmp = Path(tempfile.gettempdir()) / "fb_seed_uebersicht.xlsm"
    shutil.copy2(FB_PATH, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True, keep_vba=False)
    ws = wb["Übersicht"] if "Übersicht" in wb.sheetnames else wb[
        next(s for s in wb.sheetnames if s.replace("﻿", "").endswith("bersicht"))]
    out: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        label = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        roh = r[3] if len(r) > 3 else None
        if not label or roh is None:
            continue
        try:
            wert = float(str(roh).replace(".", "").replace(",", ".")) if isinstance(roh, str) else float(roh)
        except (ValueError, TypeError):
            continue   # z. B. '#REF!'
        out[label] = round(wert * 100)
    wb.close()
    return out


def _config_betraege() -> dict[str, int]:
    """Liest {Name: Cent} aus dem config-Blatt (Spalte B = Name, Spalte D = Betrag)."""
    tmp = Path(tempfile.gettempdir()) / "fb_seed_copy.xlsm"
    shutil.copy2(FB_PATH, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True, keep_vba=False)
    ws = wb["config"]
    out: dict[str, int] = {}
    for r in ws.iter_rows(values_only=True):
        name = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        betrag = r[3] if len(r) > 3 else None
        if name and isinstance(betrag, (int, float)):
            out[name] = round(float(betrag) * 100)
    wb.close()
    return out


def _wert_fuer(label: str, uebersicht: dict[str, int]) -> int:
    """Sucht den ersten Übersicht-Eintrag, dessen Label das Suchmuster enthält."""
    for k, v in uebersicht.items():
        if label.lower() in k.lower():
            return v
    return 0


def seed() -> None:
    betraege = _config_betraege()
    uebersicht = _uebersicht_werte()
    conn = connect()
    n_kat = n_ukat = 0
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT name, id FROM kategorien")
            kat_id = dict(cur.fetchall())

            # Kategorie-Soll
            for cname, kat in CONFIG_KATEGORIE.items():
                if cname in betraege and kat in kat_id:
                    cur.execute("UPDATE kategorien SET monatliche_ruecklage_cent=%s WHERE id=%s",
                                (betraege[cname], kat_id[kat]))
                    n_kat += 1

            # Unterkategorie-Soll (Unterkategorie ggf. anlegen)
            for cname, (kat, ukat) in CONFIG_UNTERKATEGORIE.items():
                if kat not in kat_id:
                    continue
                betrag = betraege.get(cname, 0)
                cur.execute("""INSERT INTO unterkategorien (kategorie_id, name, monatliche_ruecklage_cent, quelle)
                               VALUES (%s,%s,%s,'manuell')
                               ON CONFLICT (kategorie_id, name)
                               DO UPDATE SET monatliche_ruecklage_cent=EXCLUDED.monatliche_ruecklage_cent""",
                            (kat_id[kat], ukat, betrag))
                n_ukat += 1

            # Vermögensposten — Beträge aus der lokalen FB (nicht aus dem Code)
            for label, name, art, notiz in POSTEN_QUELLEN:
                cent = _wert_fuer(label, uebersicht) if label else 0
                cur.execute("""INSERT INTO vermoegensposten (name, wert_cent, art, notiz)
                               VALUES (%s,%s,%s,%s)
                               ON CONFLICT (name) DO UPDATE
                               SET wert_cent=EXCLUDED.wert_cent, art=EXCLUDED.art, notiz=EXCLUDED.notiz""",
                            (name, cent, art, notiz))
        conn.commit()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM vermoegensposten WHERE aktiv")
            n_posten = cur.fetchone()[0]
        print(f"[seed_uebersicht] Kategorie-Soll: {n_kat} | Unterkategorie-Soll: {n_ukat} | Vermögensposten: {n_posten}")
    finally:
        conn.close()


if __name__ == "__main__":
    seed()
